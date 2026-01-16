"""
Excel 生成模块 - 创建带甘特图的项目计划 Excel
支持计划/实际双行甘特图显示和公式关联
"""

from datetime import datetime, timedelta
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.comments import Comment

from .scheduler import Task, Scheduler
from .progress_manager import ProgressManager


class ExcelGenerator:
    """Excel 甘特图生成器"""

    def __init__(self):
        # 样式定义
        self.header_font = Font(name="微软雅黑", size=16, bold=True)
        self.title_font = Font(name="微软雅黑", size=11, bold=True)
        self.normal_font = Font(name="微软雅黑", size=10)
        self.small_font = Font(name="微软雅黑", size=8)

        # 表头填充色
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.milestone_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        # 甘特图颜色
        self.gantt_plan_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")  # 蓝色 - 计划
        self.gantt_actual_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")  # 红色 - 实际进行中
        self.gantt_complete_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")  # 绿色 - 已完成
        self.weekend_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        self.today_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        # 条件格式颜色
        self.delay_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # 延期红
        self.ontime_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # 准时绿
        self.early_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")  # 提前蓝

        # 实际行背景色（浅灰）
        self.actual_row_fill = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")

        # 进度相关颜色
        self.progress_pending_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")  # 灰色-待填写
        self.no_progress_fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")  # 橙色-无进度
        self.issue_border = Border(
            left=Side(style='medium', color='FF0000'),
            right=Side(style='medium', color='FF0000'),
            top=Side(style='medium', color='FF0000'),
            bottom=Side(style='medium', color='FF0000')
        )  # 红色边框-有问题

        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def generate(self,
                 tasks: List[Task],
                 project_name: str,
                 start_date: datetime,
                 output_path: str,
                 gantt_days: int = 90,
                 exclude_weekends: bool = True,
                 exclude_holidays: bool = False,
                 progress_manager: Optional[ProgressManager] = None,
                 gantt_start_date: Optional[datetime] = None) -> str:
        """
        生成带甘特图的 Excel 文件

        Args:
            tasks: 任务列表
            project_name: 项目名称
            start_date: 项目开始日期（用于排期计算）
            output_path: 输出文件路径
            gantt_days: 甘特图显示天数
            exclude_weekends: 是否排除周末计算工作日
            exclude_holidays: 是否排除节假日
            progress_manager: 进度管理器（可选，用于生成进度历史工作表）
            gantt_start_date: 甘特图开始日期（默认使用 start_date）

        Returns:
            生成的文件路径
        """
        # 甘特图开始日期，默认使用排期开始日期
        effective_gantt_start = gantt_start_date or start_date
        # 计算任务日期
        scheduler = Scheduler(exclude_weekends=exclude_weekends,
                             exclude_holidays=exclude_holidays)
        tasks = scheduler.calculate_dates(tasks, start_date)

        # 自动计算甘特图天数：根据任务结束日期
        if gantt_days <= 0 or gantt_days is None:
            # 找到最晚的结束日期
            end_dates = [t.end_date for t in tasks if t.end_date]
            if end_dates:
                max_end_date = max(end_dates)
                # 甘特图天数 = 最晚结束日期 - 甘特图开始日期 + 额外buffer天数
                gantt_days = (max_end_date - effective_gantt_start).days + 14  # 加14天buffer
            else:
                gantt_days = 90  # 默认90天

        wb = Workbook()
        ws = wb.active
        ws.title = "项目计划"

        # 创建标题区域（传入任务列表用于计算总天数）
        self._create_header(ws, project_name, tasks)

        # 创建表头（使用甘特图开始日期）
        self._create_table_header(ws, effective_gantt_start, gantt_days)

        # 填充任务数据（双行模式，使用甘特图开始日期）
        last_data_row = self._fill_tasks(ws, tasks, effective_gantt_start, gantt_days, progress_manager)

        # 添加条件格式（工期差异列）
        self._add_conditional_formatting(ws, last_data_row)

        # 添加图例
        self._add_legend(ws, last_data_row)

        # 冻结窗格（Q列开始是甘特图）
        ws.freeze_panes = 'Q6'

        # 如果有进度记录，创建进度历史工作表
        if progress_manager and progress_manager.records:
            self._create_progress_history_sheet(wb, tasks, progress_manager)

        # 保存文件
        wb.save(output_path)

        return output_path

    def _create_header(self, ws, project_name: str, tasks: List[Task]):
        """创建标题区域"""
        ws.merge_cells('A1:N1')
        ws['A1'] = f"{project_name}计划表"
        ws['A1'].font = self.header_font
        ws['A1'].alignment = self.center_align

        # 计算项目总天数
        total_days = 0
        if tasks:
            # 找到最早开始日期和最晚结束日期
            start_dates = [t.start_date for t in tasks if t.start_date]
            end_dates = [t.end_date for t in tasks if t.end_date]
            if start_dates and end_dates:
                project_start = min(start_dates)
                project_end = max(end_dates)
                total_days = (project_end - project_start).days + 1

        # 显示项目总天数
        ws['O1'] = f"项目总天数: {total_days}天"
        ws['O1'].font = Font(name="微软雅黑", size=11, bold=True)
        ws['O1'].alignment = Alignment(horizontal='center', vertical='center')

        ws['Q1'] = f"创建日期: {datetime.now().strftime('%Y-%m-%d')}"
        ws['Q1'].font = self.normal_font

        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 10  # 空行

    def _create_table_header(self, ws, start_date: datetime, gantt_days: int):
        """创建表头"""
        # 新列布局: A-P 为数据列，Q 起为甘特图
        # RACI 列：R-执行者、A-批准人、C-咨询人、I-知会人
        headers = ["里程碑", "编号", "任务名称",
                   "R-执行者", "A-批准人", "C-咨询人", "I-知会人",
                   "前置任务", "计划开始", "计划结束", "计划工期",
                   "实际开始", "实际结束", "进度偏差",
                   "状态", "类型"]
        header_widths = [12, 6, 32, 12, 10, 12, 12, 8, 11, 11, 8, 11, 11, 8, 8, 6]

        # 第 3-5 行为表头（3-4行合并为主表头，5行为日期星期）
        for col, (header, width) in enumerate(zip(headers, header_widths), start=1):
            # 合并 3-4 行
            ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
            ws.column_dimensions[get_column_letter(col)].width = width

        # 甘特图日期列（从 Q 列 = 17 列开始）
        gantt_start_col = 17
        today = datetime.now().date()

        for i in range(gantt_days):
            current_date = start_date + timedelta(days=i)
            col = gantt_start_col + i

            # 第 3 行: 月份（每月第一天显示）
            cell3 = ws.cell(row=3, column=col)
            if current_date.day == 1 or i == 0:
                cell3.value = f"{current_date.month}月"
            cell3.font = self.small_font
            cell3.alignment = self.center_align
            cell3.border = self.thin_border

            # 第 4 行: 日期
            cell4 = ws.cell(row=4, column=col, value=current_date.day)
            cell4.font = self.small_font
            cell4.alignment = self.center_align
            cell4.border = self.thin_border

            # 第 5 行: 星期
            weekday_names = ['一', '二', '三', '四', '五', '六', '日']
            cell5 = ws.cell(row=5, column=col, value=weekday_names[current_date.weekday()])
            cell5.font = self.small_font
            cell5.alignment = self.center_align
            cell5.border = self.thin_border

            # 周末着色
            if current_date.weekday() >= 5:
                cell3.fill = self.weekend_fill
                cell4.fill = self.weekend_fill
                cell5.fill = self.weekend_fill

            # 今天高亮
            if current_date.date() == today:
                cell3.fill = self.today_fill
                cell4.fill = self.today_fill
                cell5.fill = self.today_fill

            ws.column_dimensions[get_column_letter(col)].width = 3

        ws.row_dimensions[3].height = 18
        ws.row_dimensions[4].height = 18
        ws.row_dimensions[5].height = 16

    def _fill_tasks(self, ws, tasks: List[Task], start_date: datetime, gantt_days: int,
                    progress_manager: Optional[ProgressManager] = None) -> int:
        """
        填充任务数据（双行模式：每个任务占2行）

        Args:
            ws: 工作表
            tasks: 任务列表
            start_date: 开始日期
            gantt_days: 甘特图天数
            progress_manager: 进度管理器（可选，用于关联进度记录）

        Returns:
            最后一个数据行的行号
        """
        gantt_start_col = 17  # Q列
        current_row = 6
        current_milestone = None
        milestone_start_row = 6
        milestone_rows = []  # 记录需要合并的里程碑区域

        for task in tasks:
            plan_row = current_row
            actual_row = current_row + 1

            # 里程碑变化时，记录前一个里程碑的合并范围
            if task.milestone != current_milestone:
                if current_milestone is not None and current_row > milestone_start_row:
                    milestone_rows.append((milestone_start_row, current_row - 1))
                current_milestone = task.milestone
                milestone_start_row = current_row

            # ========== 计划行 (plan_row) ==========

            # A列: 里程碑
            cell_a = ws.cell(row=plan_row, column=1, value=task.milestone)
            cell_a.font = self.title_font
            cell_a.fill = self.milestone_fill
            cell_a.alignment = self.center_align
            cell_a.border = self.thin_border

            # B列: 编号
            cell_b = ws.cell(row=plan_row, column=2, value=task.task_no)
            cell_b.font = self.normal_font
            cell_b.alignment = self.center_align
            cell_b.border = self.thin_border

            # C列: 任务名称
            cell_c = ws.cell(row=plan_row, column=3, value=task.name)
            cell_c.font = self.normal_font
            cell_c.alignment = self.left_align
            cell_c.border = self.thin_border

            # D列: R-执行者 (RACI - Responsible)
            responsible_str = ", ".join(task.responsible) if hasattr(task, 'responsible') and task.responsible else ""
            cell_d = ws.cell(row=plan_row, column=4, value=responsible_str)
            cell_d.font = self.normal_font
            cell_d.alignment = self.center_align
            cell_d.border = self.thin_border

            # E列: A-批准人 (RACI - Accountable)
            accountable_str = task.accountable if hasattr(task, 'accountable') and task.accountable else ""
            cell_e = ws.cell(row=plan_row, column=5, value=accountable_str)
            cell_e.font = self.normal_font
            cell_e.alignment = self.center_align
            cell_e.border = self.thin_border

            # F列: C-咨询人 (RACI - Consulted)
            consulted_str = ", ".join(task.consulted) if hasattr(task, 'consulted') and task.consulted else ""
            cell_f = ws.cell(row=plan_row, column=6, value=consulted_str)
            cell_f.font = self.normal_font
            cell_f.alignment = self.center_align
            cell_f.border = self.thin_border

            # G列: I-知会人 (RACI - Informed)
            informed_str = ", ".join(task.informed) if hasattr(task, 'informed') and task.informed else ""
            cell_g = ws.cell(row=plan_row, column=7, value=informed_str)
            cell_g.font = self.normal_font
            cell_g.alignment = self.center_align
            cell_g.border = self.thin_border

            # H列: 前置任务
            cell_h = ws.cell(row=plan_row, column=8, value=task.predecessor or "-")
            cell_h.font = self.normal_font
            cell_h.alignment = self.center_align
            cell_h.border = self.thin_border

            # I列: 计划开始日期
            cell_i = ws.cell(row=plan_row, column=9)
            if task.start_date:
                cell_i.value = task.start_date
                cell_i.number_format = 'YYYY-MM-DD'
            cell_i.font = self.normal_font
            cell_i.alignment = self.center_align
            cell_i.border = self.thin_border

            # J列: 计划结束日期
            cell_j = ws.cell(row=plan_row, column=10)
            if task.end_date:
                cell_j.value = task.end_date
                cell_j.number_format = 'YYYY-MM-DD'
            cell_j.font = self.normal_font
            cell_j.alignment = self.center_align
            cell_j.border = self.thin_border

            # K列: 计划工期（公式）
            cell_k = ws.cell(row=plan_row, column=11)
            cell_k.value = f'=IF(AND(I{plan_row}<>"",J{plan_row}<>""),J{plan_row}-I{plan_row}+1,"")'
            cell_k.font = self.normal_font
            cell_k.alignment = self.center_align
            cell_k.border = self.thin_border

            # L列: 实际开始日期
            cell_l = ws.cell(row=plan_row, column=12)
            if task.actual_start:
                cell_l.value = task.actual_start
                cell_l.number_format = 'YYYY-MM-DD'
            cell_l.font = self.normal_font
            cell_l.alignment = self.center_align
            cell_l.border = self.thin_border

            # M列: 实际结束日期
            cell_m = ws.cell(row=plan_row, column=13)
            if task.actual_end:
                cell_m.value = task.actual_end
                cell_m.number_format = 'YYYY-MM-DD'
            cell_m.font = self.normal_font
            cell_m.alignment = self.center_align
            cell_m.border = self.thin_border

            # N列: 进度偏差（公式：实际结束 - 计划结束，正数延期，负数提前）
            cell_n = ws.cell(row=plan_row, column=14)
            cell_n.value = f'=IF(AND(M{plan_row}<>"",J{plan_row}<>""),M{plan_row}-J{plan_row},"")'
            cell_n.font = self.normal_font
            cell_n.alignment = self.center_align
            cell_n.border = self.thin_border

            # O列: 状态（公式）
            cell_o = ws.cell(row=plan_row, column=15)
            cell_o.value = f'=IF(M{plan_row}<>"","已完成",IF(L{plan_row}<>"","进行中","未开始"))'
            cell_o.font = self.normal_font
            cell_o.alignment = self.center_align
            cell_o.border = self.thin_border

            # P列: 类型标识 - 计划
            cell_p = ws.cell(row=plan_row, column=16, value="计划")
            cell_p.font = Font(name="微软雅黑", size=9, color="5B9BD5")
            cell_p.alignment = self.center_align
            cell_p.border = self.thin_border

            # ========== 实际行 (actual_row) ==========

            # A-O列: 实际行空白（这些列在实际行留空或使用浅色背景）
            for col in range(1, 16):
                cell = ws.cell(row=actual_row, column=col)
                cell.fill = self.actual_row_fill
                cell.border = self.thin_border

            # P列: 类型标识 - 实际
            cell_p2 = ws.cell(row=actual_row, column=16, value="实际")
            cell_p2.font = Font(name="微软雅黑", size=9, color="E74C3C")
            cell_p2.alignment = self.center_align
            cell_p2.border = self.thin_border
            cell_p2.fill = self.actual_row_fill

            # ========== 绘制甘特图条 ==========

            for i in range(gantt_days):
                current_date = start_date + timedelta(days=i)
                col = gantt_start_col + i

                # 计划行甘特图
                cell_plan = ws.cell(row=plan_row, column=col)
                cell_plan.border = self.thin_border

                in_plan = (task.start_date and task.end_date and
                          task.start_date.date() <= current_date.date() <= task.end_date.date())

                if in_plan:
                    cell_plan.fill = self.gantt_plan_fill
                elif current_date.weekday() >= 5:
                    cell_plan.fill = self.weekend_fill

                # 实际行甘特图（含进度记录关联）
                cell_actual = ws.cell(row=actual_row, column=col)
                cell_actual.border = self.thin_border

                # 查找当天的进度记录
                day_record = self._find_record_for_date(progress_manager, task.task_no, current_date)

                # 判断当前日期是否在计划范围内
                in_plan_range = (task.start_date and task.end_date and
                                task.start_date.date() <= current_date.date() <= task.end_date.date())

                # 实际行填充逻辑：
                # 1. 有进度记录且增量 > 0 → 绿色 + 显示增量
                # 2. 有进度记录但增量 <= 0（无进度）→ 橙色 + 显示"0"
                # 3. 有问题记录 → 红色边框
                # 4. 在计划范围内但无记录（漏填）→ 浅灰色
                # 5. 周末 → 周末色
                # 6. 其他 → 默认背景

                if day_record:
                    # 有当天的进度记录，获取增量
                    increment = getattr(day_record, 'increment', 0)

                    if increment > 0:
                        # 有进度增量：绿色
                        cell_actual.fill = self.gantt_complete_fill
                        cell_actual.font = Font(name="微软雅黑", size=7, color="FFFFFF", bold=True)
                        cell_actual.value = f"+{increment}%"
                    else:
                        # 无进度增量（增量 <= 0）：橙色
                        cell_actual.fill = self.no_progress_fill
                        cell_actual.font = Font(name="微软雅黑", size=7, color="FFFFFF", bold=True)
                        cell_actual.value = f"{increment}%" if increment < 0 else "0"

                    cell_actual.alignment = self.center_align

                    # 添加批注
                    comment_lines = []
                    record_date_str = day_record.record_date.strftime('%Y-%m-%d') if hasattr(day_record.record_date, 'strftime') else str(day_record.record_date)
                    comment_lines.append(f"日期: {record_date_str}")
                    comment_lines.append(f"累计进度: {day_record.progress}%")
                    comment_lines.append(f"当日增量: {'+' if increment > 0 else ''}{increment}%")
                    comment_lines.append(f"状态: {day_record.status.value if hasattr(day_record.status, 'value') else day_record.status}")
                    if day_record.note:
                        comment_lines.append(f"备注: {day_record.note}")
                    if day_record.issues:
                        # 无进度原因或问题
                        comment_lines.append(f"问题/原因: {day_record.issues}")

                    comment_text = "\n".join(comment_lines)
                    cell_actual.comment = Comment(comment_text, "APQP系统")
                    cell_actual.comment.width = 250
                    cell_actual.comment.height = 100

                    # 有问题时添加红色边框
                    if day_record.issues:
                        cell_actual.border = self.issue_border

                elif in_plan_range:
                    # 在计划范围内但无进度记录：浅灰色（漏填/待填写）
                    cell_actual.fill = self.progress_pending_fill
                elif current_date.weekday() >= 5:
                    # 周末
                    cell_actual.fill = self.weekend_fill
                else:
                    # 其他：默认背景
                    cell_actual.fill = self.actual_row_fill

            ws.row_dimensions[plan_row].height = 22
            ws.row_dimensions[actual_row].height = 18

            current_row += 2  # 每个任务占2行

        # 处理最后一个里程碑的合并
        if current_row > milestone_start_row:
            milestone_rows.append((milestone_start_row, current_row - 1))

        # 合并里程碑单元格
        for start_r, end_r in milestone_rows:
            if end_r > start_r:
                ws.merge_cells(f'A{start_r}:A{end_r}')

        # 添加进度记录详情列（甘特图最右侧）
        if progress_manager and progress_manager.records:
            progress_col = gantt_start_col + gantt_days

            # 添加表头
            ws.merge_cells(start_row=3, start_column=progress_col, end_row=5, end_column=progress_col)
            header_cell = ws.cell(row=3, column=progress_col, value="进度记录")
            header_cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
            header_cell.fill = self.header_fill
            header_cell.alignment = self.center_align
            header_cell.border = self.thin_border
            ws.column_dimensions[get_column_letter(progress_col)].width = 28

            # 填充每个任务的最近进度记录
            task_row = 6
            for task in tasks:
                recent_records = self._get_recent_records_summary(progress_manager, task.task_no, limit=3)
                if recent_records:
                    summary_lines = []
                    for r in recent_records:
                        date_str = r.record_date.strftime('%m-%d') if hasattr(r.record_date, 'strftime') else str(r.record_date)[:5]
                        note_preview = r.note[:8] + "..." if r.note and len(r.note) > 8 else (r.note or "")
                        summary_lines.append(f"{date_str}: {r.progress}% {note_preview}")

                    summary_text = "\n".join(summary_lines)
                else:
                    summary_text = ""

                # 合并计划行和实际行的进度记录列
                ws.merge_cells(start_row=task_row, start_column=progress_col,
                              end_row=task_row + 1, end_column=progress_col)
                summary_cell = ws.cell(row=task_row, column=progress_col, value=summary_text)
                summary_cell.font = Font(name="微软雅黑", size=8)
                summary_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                summary_cell.border = self.thin_border

                task_row += 2

        return current_row - 1

    def _add_conditional_formatting(self, ws, last_row: int):
        """添加条件格式（进度偏差列）"""
        # N列进度偏差的条件格式（正数延期红，负数提前蓝，0准时绿）
        range_str = f'N6:N{last_row}'

        # 延期（>0）- 红色
        ws.conditional_formatting.add(
            range_str,
            CellIsRule(operator='greaterThan', formula=['0'], fill=self.delay_fill)
        )

        # 准时（=0）- 绿色
        ws.conditional_formatting.add(
            range_str,
            CellIsRule(operator='equal', formula=['0'], fill=self.ontime_fill)
        )

        # 提前（<0）- 蓝色
        ws.conditional_formatting.add(
            range_str,
            CellIsRule(operator='lessThan', formula=['0'], fill=self.early_fill)
        )

    def _add_legend(self, ws, last_row: int):
        """添加图例说明"""
        legend_row = last_row + 3

        ws.cell(row=legend_row, column=1, value="甘特图:").font = self.title_font

        # 计划 - 蓝色
        cell1 = ws.cell(row=legend_row, column=3)
        cell1.fill = self.gantt_plan_fill
        ws.cell(row=legend_row, column=4, value="计划进度").font = self.normal_font

        # 已完成部分 - 绿色
        cell2 = ws.cell(row=legend_row, column=6)
        cell2.fill = self.gantt_complete_fill
        ws.cell(row=legend_row, column=7, value="已完成部分").font = self.normal_font

        # 未完成部分 - 灰色
        cell3 = ws.cell(row=legend_row, column=9)
        cell3.fill = self.progress_pending_fill
        ws.cell(row=legend_row, column=10, value="未完成部分").font = self.normal_font

        # 进行中 - 红色
        cell4 = ws.cell(row=legend_row, column=12)
        cell4.fill = self.gantt_actual_fill
        ws.cell(row=legend_row, column=13, value="进行中").font = self.normal_font

        # 进度偏差说明（实际结束 - 计划结束）
        legend_row2 = legend_row + 1
        ws.cell(row=legend_row2, column=1, value="进度偏差:").font = self.title_font

        cell4 = ws.cell(row=legend_row2, column=3)
        cell4.fill = self.delay_fill
        ws.cell(row=legend_row2, column=4, value=">0 延期").font = self.normal_font

        cell5 = ws.cell(row=legend_row2, column=6)
        cell5.fill = self.ontime_fill
        ws.cell(row=legend_row2, column=7, value="=0 准时").font = self.normal_font

        cell6 = ws.cell(row=legend_row2, column=9)
        cell6.fill = self.early_fill
        ws.cell(row=legend_row2, column=10, value="<0 提前").font = self.normal_font

    def _find_record_for_date(self, progress_manager: Optional[ProgressManager],
                               task_no: str, date: datetime) -> Optional[dict]:
        """
        查找指定任务在指定日期的进度记录

        Args:
            progress_manager: 进度管理器
            task_no: 任务编号
            date: 查询日期

        Returns:
            进度记录字典或 None
        """
        if not progress_manager:
            return None

        for record in progress_manager.records.values():
            if record.task_no == task_no:
                record_date = record.record_date
                if hasattr(record_date, 'date'):
                    record_date = record_date.date()
                target_date = date.date() if hasattr(date, 'date') else date
                if record_date == target_date:
                    return record
        return None

    def _get_recent_records_summary(self, progress_manager: Optional[ProgressManager],
                                     task_no: str, limit: int = 3) -> list:
        """
        获取任务最近N条记录摘要

        Args:
            progress_manager: 进度管理器
            task_no: 任务编号
            limit: 返回记录数量

        Returns:
            最近记录列表
        """
        if not progress_manager:
            return []

        task_records = [r for r in progress_manager.records.values() if r.task_no == task_no]
        task_records.sort(key=lambda r: r.record_date, reverse=True)
        return task_records[:limit]

    def _create_progress_history_sheet(self, wb: Workbook, tasks: List[Task],
                                       progress_manager: ProgressManager):
        """创建进度历史工作表"""
        ws = wb.create_sheet("进度历史")

        # 创建任务名称查找字典
        task_names = {task.task_no: task.name for task in tasks}

        # 表头
        headers = ["任务编号", "任务名称", "记录日期", "完成进度", "状态", "备注", "问题"]
        header_widths = [10, 30, 12, 10, 10, 30, 30]

        for col, (header, width) in enumerate(zip(headers, header_widths), start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 25

        # 获取所有进度记录并按日期排序
        all_records = list(progress_manager.records.values())
        all_records.sort(key=lambda r: (r.record_date, r.task_no))

        # 填充数据
        for row, record in enumerate(all_records, start=2):
            # 任务编号
            cell_a = ws.cell(row=row, column=1, value=record.task_no)
            cell_a.font = self.normal_font
            cell_a.alignment = self.center_align
            cell_a.border = self.thin_border

            # 任务名称
            task_name = task_names.get(record.task_no, "未知任务")
            cell_b = ws.cell(row=row, column=2, value=task_name)
            cell_b.font = self.normal_font
            cell_b.alignment = self.left_align
            cell_b.border = self.thin_border

            # 记录日期
            cell_c = ws.cell(row=row, column=3, value=record.record_date)
            cell_c.number_format = 'YYYY-MM-DD'
            cell_c.font = self.normal_font
            cell_c.alignment = self.center_align
            cell_c.border = self.thin_border

            # 完成进度
            cell_d = ws.cell(row=row, column=4, value=f"{record.progress}%")
            cell_d.font = self.normal_font
            cell_d.alignment = self.center_align
            cell_d.border = self.thin_border

            # 状态
            cell_e = ws.cell(row=row, column=5, value=record.status.value)
            cell_e.font = self.normal_font
            cell_e.alignment = self.center_align
            cell_e.border = self.thin_border

            # 备注
            cell_f = ws.cell(row=row, column=6, value=record.note or "")
            cell_f.font = self.normal_font
            cell_f.alignment = self.left_align
            cell_f.border = self.thin_border

            # 问题
            cell_g = ws.cell(row=row, column=7, value=record.issues or "")
            cell_g.font = self.normal_font
            cell_g.alignment = self.left_align
            cell_g.border = self.thin_border

            ws.row_dimensions[row].height = 22

        # 冻结首行
        ws.freeze_panes = 'A2'


def generate_excel(tasks: List[Task],
                   project_name: str,
                   start_date: datetime,
                   output_path: str,
                   gantt_days: int = 90,
                   exclude_weekends: bool = True,
                   exclude_holidays: bool = False,
                   progress_manager: Optional[ProgressManager] = None) -> str:
    """
    便捷函数：生成 Excel 文件

    Args:
        tasks: 任务列表
        project_name: 项目名称
        start_date: 开始日期
        output_path: 输出路径
        gantt_days: 甘特图天数
        exclude_weekends: 排除周末
        exclude_holidays: 排除节假日
        progress_manager: 进度管理器（可选）

    Returns:
        生成的文件路径
    """
    generator = ExcelGenerator()
    return generator.generate(
        tasks=tasks,
        project_name=project_name,
        start_date=start_date,
        output_path=output_path,
        gantt_days=gantt_days,
        exclude_weekends=exclude_weekends,
        exclude_holidays=exclude_holidays,
        progress_manager=progress_manager
    )
