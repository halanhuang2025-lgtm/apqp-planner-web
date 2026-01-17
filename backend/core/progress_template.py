"""
批量进度模板生成器 - 用于跨项目批量导入进度
"""

from datetime import datetime
from typing import List, Optional, Dict, Tuple
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.utils import get_column_letter

from .scheduler import Task, TaskStatus
from .project_manager import ProjectManager


class BatchProgressTemplateGenerator:
    """批量进度模板生成器"""

    def __init__(self):
        # 样式定义
        self.header_font = Font(name="微软雅黑", size=11, bold=True, color="000000")
        self.normal_font = Font(name="微软雅黑", size=10)
        self.project_font = Font(name="微软雅黑", size=10, bold=True)

        # 表头填充色
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.project_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        self.editable_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # 黄色-可编辑
        self.overdue_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")   # 浅红色-已过期

        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def generate(self,
                 project_manager: ProjectManager,
                 output_path: str,
                 record_date: Optional[datetime] = None) -> str:
        """
        生成批量进度模板 Excel

        Args:
            project_manager: 项目管理器
            output_path: 输出文件路径
            record_date: 记录日期（默认今天）

        Returns:
            生成的文件路径
        """
        record_date = record_date or datetime.now()

        wb = Workbook()
        ws = wb.active
        ws.title = "进度导入模板"

        # 创建表头
        self._create_header(ws, record_date)

        # 获取所有活跃项目并填充任务
        projects = project_manager.list_projects(status="active")
        row = 4  # 数据从第4行开始（第1行标题，第2行提示，第3行列标题）
        is_first_project = True

        for project in projects:
            tasks, progress_manager, _ = project_manager.load_project_data(project.id)
            if not tasks:
                continue

            # 非排除的任务
            valid_tasks = [t for t in tasks if not t.excluded]
            if not valid_tasks:
                continue

            # 项目之间添加空行分隔（第一个项目除外）
            if not is_first_project:
                row += 1  # 空一行
            is_first_project = False

            # 填充项目任务
            for task in valid_tasks:
                self._fill_task_row(ws, row, project, task, record_date)
                row += 1

        # 设置列宽
        self._set_column_widths(ws)

        # 保护工作表（只允许编辑可编辑列，不设密码）
        ws.protection.sheet = True
        ws.protection.enable()

        # 保存文件
        wb.save(output_path)
        return output_path

    def _create_header(self, ws, record_date: datetime):
        """创建表头"""
        # 第一行：标题
        ws.merge_cells('A1:N1')
        ws['A1'] = f"每日进度批量导入模板 - 记录日期: {record_date.strftime('%Y-%m-%d')}"
        ws['A1'].font = Font(name="微软雅黑", size=12, bold=True)
        ws['A1'].alignment = self.center_align
        ws.row_dimensions[1].height = 30

        # 第二行：填写提示
        ws.merge_cells('A2:N2')
        ws['A2'] = "【填写说明】请执行人(R)在黄色列填写：「今日增加进度」输入百分比如10%；「今日备注」填当日完成的具体工作；「问题/异常」填遇到的困难、风险或需要协调的事项"
        ws['A2'].font = Font(name="微软雅黑", size=9, italic=True, color="666666")
        ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[2].height = 22

        # 第三行：列标题
        headers = [
            ("项目名称", 35, False),
            ("项目编号", 12, False),
            ("项目分类", 14, False),       # 新增：新产品开发、特殊定制、工程项目非标机
            ("整机编码", 18, False),
            ("任务编号", 10, False),
            ("任务名称", 30, False),
            ("里程碑", 15, False),
            ("执行人(R)", 12, False),
            ("计划开始", 12, False),
            ("计划结束", 12, False),
            ("当前进度", 10, False),
            ("今日增加进度", 14, True),   # 可编辑，百分比
            ("今日备注", 25, True),   # 可编辑
            ("问题/异常", 25, True),  # 可编辑
        ]

        for col, (header, width, editable) in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill if not editable else self.editable_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[3].height = 25

    def _fill_task_row(self, ws, row: int, project, task: Task, record_date: datetime):
        """填充任务行"""
        # 检查是否过期：计划开始日期已到或已过 且 进度未完成
        is_overdue = False
        if task.start_date and task.progress < 100:
            if task.start_date.date() <= record_date.date():
                is_overdue = True

        # A: 项目名称 (只读)
        cell = ws.cell(row=row, column=1, value=project.name)
        cell.font = self.project_font
        cell.fill = self.overdue_fill if is_overdue else self.project_fill
        cell.alignment = self.left_align
        cell.border = self.thin_border
        cell.protection = Protection(locked=True)

        # B: 项目编号 (只读)
        cell = ws.cell(row=row, column=2, value=project.project_no or "")
        cell.font = self.normal_font
        if is_overdue:
            cell.fill = self.overdue_fill
        cell.alignment = self.center_align
        cell.border = self.thin_border
        cell.protection = Protection(locked=True)

        # C: 项目分类 (只读)
        cell = ws.cell(row=row, column=3, value=project.project_type or "")
        cell.font = self.normal_font
        if is_overdue:
            cell.fill = self.overdue_fill
        cell.alignment = self.center_align
        cell.border = self.thin_border
        cell.protection = Protection(locked=True)

        # D: 整机编码 (只读)
        cell = ws.cell(row=row, column=4, value=project.machine_no or "")
        cell.font = self.normal_font
        if is_overdue:
            cell.fill = self.overdue_fill
        cell.alignment = self.center_align
        cell.border = self.thin_border
        cell.protection = Protection(locked=True)

        # E: 任务编号 (只读)
        cell = ws.cell(row=row, column=5, value=task.task_no)
        cell.font = self.normal_font
        if is_overdue:
            cell.fill = self.overdue_fill
        cell.alignment = self.center_align
        cell.border = self.thin_border
        cell.protection = Protection(locked=True)

        # F: 任务名称 (只读)
        cell = ws.cell(row=row, column=6, value=task.name)
        cell.font = self.normal_font
        if is_overdue:
            cell.fill = self.overdue_fill
        cell.alignment = self.left_align
        cell.border = self.thin_border
        cell.protection = Protection(locked=True)

        # G: 里程碑 (只读)
        cell = ws.cell(row=row, column=7, value=task.milestone)
        cell.font = self.normal_font
        if is_overdue:
            cell.fill = self.overdue_fill
        cell.alignment = self.center_align
        cell.border = self.thin_border
        cell.protection = Protection(locked=True)

        # H: 执行人(R) (只读) - 显示RACI中的R
        responsible_str = ", ".join(task.responsible) if task.responsible else ""
        cell = ws.cell(row=row, column=8, value=responsible_str)
        cell.font = self.normal_font
        if is_overdue:
            cell.fill = self.overdue_fill
        cell.alignment = self.center_align
        cell.border = self.thin_border
        cell.protection = Protection(locked=True)

        # I: 计划开始 (只读) - 过期时高亮显示
        cell = ws.cell(row=row, column=9)
        if task.start_date:
            cell.value = task.start_date.strftime("%Y-%m-%d")
        cell.font = self.normal_font
        if is_overdue:
            cell.fill = self.overdue_fill
        cell.alignment = self.center_align
        cell.border = self.thin_border
        cell.protection = Protection(locked=True)

        # J: 计划结束 (只读)
        cell = ws.cell(row=row, column=10)
        if task.end_date:
            cell.value = task.end_date.strftime("%Y-%m-%d")
        cell.font = self.normal_font
        if is_overdue:
            cell.fill = self.overdue_fill
        cell.alignment = self.center_align
        cell.border = self.thin_border
        cell.protection = Protection(locked=True)

        # K: 当前进度 (只读)
        cell = ws.cell(row=row, column=11, value=f"{task.progress}%")
        cell.font = self.normal_font
        if is_overdue:
            cell.fill = self.overdue_fill
        cell.alignment = self.center_align
        cell.border = self.thin_border
        cell.protection = Protection(locked=True)

        # L: 今日增加进度 (可编辑) - 黄色背景，百分比格式
        cell = ws.cell(row=row, column=12, value="")
        cell.font = self.normal_font
        cell.fill = self.editable_fill
        cell.alignment = self.center_align
        cell.border = self.thin_border
        cell.protection = Protection(locked=False)
        cell.number_format = '0%'  # 百分比格式

        # M: 今日备注 (可编辑) - 黄色背景
        cell = ws.cell(row=row, column=13, value="")
        cell.font = self.normal_font
        cell.fill = self.editable_fill
        cell.alignment = self.left_align
        cell.border = self.thin_border
        cell.protection = Protection(locked=False)

        # N: 问题/异常 (可编辑) - 黄色背景
        cell = ws.cell(row=row, column=14, value="")
        cell.font = self.normal_font
        cell.fill = self.editable_fill
        cell.alignment = self.left_align
        cell.border = self.thin_border
        cell.protection = Protection(locked=False)

        ws.row_dimensions[row].height = 22

    def _set_column_widths(self, ws):
        """设置列宽"""
        # A:项目名称, B:项目编号, C:项目分类, D:整机编码, E:任务编号, F:任务名称, G:里程碑, H:执行人, I:计划开始, J:计划结束, K:当前进度, L:今日增加, M:备注, N:问题
        widths = [35, 12, 14, 18, 10, 30, 15, 12, 12, 12, 10, 14, 25, 25]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width


class BatchProgressImporter:
    """批量进度导入器"""

    def parse_template(self, file_path: str) -> Tuple[List[Dict], List[str]]:
        """
        解析进度模板文件

        Args:
            file_path: Excel 文件路径

        Returns:
            (解析的进度数据列表, 错误列表)
        """
        errors = []
        progress_data = []

        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
        except Exception as e:
            return [], [f"无法打开文件: {str(e)}"]

        # 从第4行开始读取数据（第1行标题，第2行提示，第3行列标题）
        # 列顺序：A项目名称, B项目编号, C项目分类, D整机编码, E任务编号, F任务名称, G里程碑, H执行人, I计划开始, J计划结束, K当前进度, L今日增加, M备注, N问题
        for row in range(4, ws.max_row + 1):
            project_name = ws.cell(row=row, column=1).value
            task_no = ws.cell(row=row, column=5).value                # E列
            current_progress_str = ws.cell(row=row, column=11).value  # K列
            increment_value = ws.cell(row=row, column=12).value       # L列
            note = ws.cell(row=row, column=13).value                  # M列
            issues = ws.cell(row=row, column=14).value                # N列

            # 跳过空行
            if not project_name or not task_no:
                continue

            # 跳过没有填写增量的行
            if increment_value is None or increment_value == "":
                continue

            # 解析今日增加进度的值（支持百分比格式和整数格式）
            # Excel 百分比格式：1 = 100%, 0.1 = 10%
            # 直接输入数字：10 = 10%, 100 = 100%
            try:
                num_value = float(increment_value)
                if -1 <= num_value <= 1:
                    # 百分比格式存储为小数，如 100% = 1, 10% = 0.1
                    increment = int(round(num_value * 100))
                else:
                    # 大于1的数值当作直接百分比处理，如 10 = 10%
                    increment = int(num_value)
            except (ValueError, TypeError):
                errors.append(f"第 {row} 行：今日增加进度值无效 '{increment_value}'")
                continue

            # 解析当前进度
            current_progress = 0
            if current_progress_str:
                try:
                    # 移除百分号
                    progress_str = str(current_progress_str).replace("%", "").strip()
                    current_progress = int(float(progress_str))
                except (ValueError, TypeError):
                    pass

            # 计算新进度
            new_progress = min(100, max(0, current_progress + increment))

            # 自动判断状态
            if new_progress == 0:
                status = "未开始"
            elif new_progress >= 100:
                status = "已完成"
            else:
                status = "进行中"

            progress_data.append({
                "project_name": str(project_name).strip(),
                "task_no": str(task_no).strip(),
                "current_progress": current_progress,
                "increment": increment,
                "new_progress": new_progress,
                "status": status,
                "note": str(note).strip() if note else "",
                "issues": str(issues).strip() if issues else "",
            })

        return progress_data, errors

    def import_progress(self,
                        project_manager: ProjectManager,
                        progress_data: List[Dict],
                        record_date: Optional[datetime] = None) -> Dict:
        """
        批量导入进度数据

        Args:
            project_manager: 项目管理器
            progress_data: 解析的进度数据
            record_date: 记录日期

        Returns:
            导入结果统计
        """
        from .progress_manager import ProgressManager

        record_date = record_date or datetime.now()
        status_map = {
            "未开始": TaskStatus.NOT_STARTED,
            "进行中": TaskStatus.IN_PROGRESS,
            "已完成": TaskStatus.COMPLETED,
            "暂停": TaskStatus.PAUSED,
        }

        # 按项目分组
        by_project: Dict[str, List[Dict]] = {}
        for item in progress_data:
            project_name = item["project_name"]
            if project_name not in by_project:
                by_project[project_name] = []
            by_project[project_name].append(item)

        # 获取所有项目的映射
        all_projects = project_manager.list_projects(status="active")
        project_map = {p.name: p for p in all_projects}

        results = {
            "success": True,
            "projects_updated": 0,
            "imported_count": 0,
            "skipped_count": 0,
            "errors": [],
            "details": [],  # 每个项目的导入详情
        }

        # 逐项目处理
        for project_name, items in by_project.items():
            project = project_map.get(project_name)
            if not project:
                results["errors"].append(f"项目不存在: {project_name}")
                results["skipped_count"] += len(items)
                continue

            # 加载项目数据
            tasks, progress_manager, milestones = project_manager.load_project_data(project.id)
            task_map = {t.task_no: (i, t) for i, t in enumerate(tasks)}

            project_imported = 0
            project_skipped = 0

            for item in items:
                task_no = item["task_no"]
                if task_no not in task_map:
                    results["errors"].append(f"[{project_name}] 任务不存在: {task_no}")
                    project_skipped += 1
                    continue

                idx, task = task_map[task_no]

                # 添加进度记录
                status = status_map.get(item["status"], TaskStatus.IN_PROGRESS)
                record = progress_manager.add_record(
                    task=task,
                    progress=item["new_progress"],
                    status=status,
                    note=item["note"],
                    issues=item["issues"],
                    record_date=record_date
                )

                # 更新任务状态
                task.progress = item["new_progress"]
                task.status = status

                # 自动更新实际日期
                if item["new_progress"] > 0 and not task.actual_start:
                    task.actual_start = record_date
                if item["new_progress"] == 100 and not task.actual_end:
                    task.actual_end = record_date

                project_imported += 1

            # 保存项目数据
            if project_imported > 0:
                project_manager.save_project_data(project.id, tasks, progress_manager, milestones)
                results["projects_updated"] += 1

            results["imported_count"] += project_imported
            results["skipped_count"] += project_skipped
            results["details"].append({
                "project_name": project_name,
                "imported": project_imported,
                "skipped": project_skipped,
            })

        return results
